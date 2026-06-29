import os
import json
import logging
import asyncio
from datetime import datetime
import openpyxl
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

from google.antigravity import Agent, LocalAgentConfig
from google.antigravity.triggers import every, TriggerContext

# Logging Setup
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("ingest_agent.log", encoding="utf-8"),
        logging.StreamHandler()
    ]
)

# API Permissions Scopes
SCOPES = [
    'https://www.googleapis.com/auth/gmail.modify',  # Read and mark as read
    'https://www.googleapis.com/auth/drive.readonly'  # Read files from Drive
]

# Google Drive folder ID to watch (set this value)
DRIVE_FOLDER_ID = "YOUR_GOOGLE_DRIVE_FOLDER_ID_HERE"

# Ensure data directory exists
os.makedirs("data", exist_ok=True)

# Helper to get Google API credentials
def get_credentials():
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception as e:
                logging.error(f"Failed to refresh token: {e}")
                creds = None
        if not creds:
            if not os.path.exists('credentials.json'):
                logging.error("Missing 'credentials.json'! Please download desktop OAuth client secrets from Google Cloud Console.")
                return None
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return creds

# Helper to load/save helper states (e.g. processed file IDs)
def get_processed_files():
    path = "data/processed_files.json"
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def save_processed_file(file_id):
    path = "data/processed_files.json"
    processed = get_processed_files()
    if file_id not in processed:
        processed.append(file_id)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(processed, f, ensure_ascii=False, indent=2)

# =============================================================================
# 1. Gmail Order Fetcher
# =============================================================================
async def check_gmail_orders(ctx: TriggerContext):
    """Checks Gmail for new unread order emails, parses attached Excel files,
    and updates data/mail_pedidos.json.
    """
    logging.info("Starting Gmail check for new sales orders...")
    creds = get_credentials()
    if not creds:
        await ctx.send("Gmail check skipped: Google API credentials missing.")
        return

    try:
        service = build('gmail', 'v1', credentials=creds)
        # Search unread emails with excel attachments having 'pedido' or 'order' in subject/body
        query = 'is:unread has:attachment filename:xlsx (subject:pedido OR subject:order OR subject:销售订单)'
        results = service.users().messages().list(userId='me', q=query).execute()
        messages = results.get('messages', [])

        if not messages:
            logging.info("No new unread order emails found.")
            return

        orders_parsed = 0
        new_records = []

        for msg in messages:
            msg_id = msg['id']
            message = service.users().messages().get(userId='me', id=msg_id).execute()
            
            # Read metadata
            headers = message['payload']['headers']
            subject = next((h['value'] for h in headers if h['name'].lower() == 'subject'), 'No Subject')
            sender = next((h['value'] for h in headers if h['name'].lower() == 'from'), 'Unknown Sender')
            date_str = next((h['value'] for h in headers if h['name'].lower() == 'date'), '')

            logging.info(f"Processing email: '{subject}' from {sender}")

            # Check attachments
            parts = message['payload'].get('parts', [])
            for part in parts:
                if part['filename'] and part['filename'].endswith('.xlsx'):
                    attachment_id = part['body'].get('attachmentId')
                    if attachment_id:
                        attachment = service.users().messages().attachments().get(
                            userId='me', messageId=msg_id, id=attachment_id
                        ).execute()
                        
                        # Decode attachment bytes
                        import base64
                        file_data = base64.urlsafe_b64decode(attachment['data'].encode('UTF-8'))
                        
                        # Save temporarily
                        temp_path = f"data/temp_gmail_{msg_id}.xlsx"
                        with open(temp_path, 'wb') as f:
                            f.write(file_data)
                        
                        # Parse Excel sheet
                        records = parse_orders_excel(temp_path, msg_id, sender, date_str)
                        if records:
                            new_records.extend(records)
                            orders_parsed += len(records)
                        
                        # Remove temp file
                        if os.path.exists(temp_path):
                            os.remove(temp_path)

            # Mark email as read by removing 'UNREAD' label
            service.users().messages().batchModify(
                userId='me',
                body={
                    'ids': [msg_id],
                    'removeLabelIds': ['UNREAD']
                }
            ).execute()

        if new_records:
            save_records_to_json(new_records, "data/mail_pedidos.json")
            await ctx.send(f"📊 Gmail 模块：成功读取并解析了 {len(new_records)} 条订单交易行，并更新了本地数据库。")

    except HttpError as error:
        logging.error(f"Gmail API Error: {error}")
        await ctx.send(f"Gmail API 出现错误: {error}")

def parse_orders_excel(file_path, msg_id, sender, date_str):
    """Parses orders from a transaction Excel template."""
    records = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        # Determine index map (assuming row 1 or 2 is header)
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        
        # Mappings
        mappings = {
            'sku': ['sku', 'sku_producto', '产品sku', '条码', '编码', 'producto_sku'],
            'qty': ['cantidad', '数量', '件数', 'qty', 'quantity'],
            'price': ['precio', '单ay', 'price', '单价', 'precio_unitario'],
            'total': ['total', '总价', '金额', 'total_linea']
        }
        
        idx_map = {}
        for key, aliases in mappings.items():
            idx_map[key] = next((i for i, h in enumerate(headers) if any(a in h.lower() for a in aliases)), -1)

        # Iterate rows
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            sku = row[idx_map['sku']].value if idx_map['sku'] != -1 else None
            qty = row[idx_map['qty']].value if idx_map['qty'] != -1 else 1
            price = row[idx_map['price']].value if idx_map['price'] != -1 else 0
            total = row[idx_map['total']].value if idx_map['total'] != -1 else 0

            if not sku:
                continue

            records.append({
                "order_id": f"GMAIL_{msg_id}_{row_idx}",
                "fecha_pedido": date_str,
                "cliente_email": sender,
                "producto_sku": str(sku).strip(),
                "cantidad": int(qty) if qty else 1,
                "precio_unitario": float(price) if price else 0.0,
                "total_linea": float(total) if total else (float(price) * int(qty) if price and qty else 0.0)
            })
        
        wb.close()
    except Exception as e:
        logging.error(f"Error parsing orders Excel: {e}")
    return records

# =============================================================================
# 2. Google Drive Power BI Watcher
# =============================================================================
async def check_gdrive_reports(ctx: TriggerContext):
    """Checks the watched Google Drive folder for newly uploaded Power BI reports,
    downloads and parses them, updating data/powerbi_estadisticas.json.
    """
    logging.info("Starting Google Drive check for Power BI reports...")
    creds = get_credentials()
    if not creds:
        await ctx.send("Drive check skipped: Google API credentials missing.")
        return

    if DRIVE_FOLDER_ID == "YOUR_GOOGLE_DRIVE_FOLDER_ID_HERE":
        logging.warning("Google Drive watching skipped: DRIVE_FOLDER_ID not set in ingest_agent.py.")
        return

    try:
        service = build('drive', 'v3', credentials=creds)
        # Query files in folder ending with xlsx/xls/csv
        query = f"'{DRIVE_FOLDER_ID}' in parents and (mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or mimeType = 'text/csv') and trashed = false"
        results = service.files().list(q=query, fields="files(id, name, modifiedTime)").execute()
        files = results.get('files', [])

        processed_ids = get_processed_files()
        new_files_count = 0
        new_records = []

        for f in files:
            file_id = f['id']
            file_name = f['name']
            
            if file_id not in processed_ids:
                logging.info(f"New Power BI report detected on Drive: {file_name}")
                new_files_count += 1
                
                # Download
                request = service.files().get_media(fileId=file_id)
                import io
                from googleapiclient.http import MediaIoBaseDownload
                
                fh = io.BytesIO()
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while not done:
                    status, done = downloader.next_chunk()
                
                # Save locally
                temp_path = f"data/temp_drive_{file_id}.xlsx"
                with open(temp_path, "wb") as out_file:
                    out_file.write(fh.getvalue())
                
                # Parse
                records = parse_powerbi_excel(temp_path)
                if records:
                    new_records.extend(records)
                
                # Clean up and register as processed
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                
                save_processed_file(file_id)

        if new_records:
            save_records_to_json(new_records, "data/powerbi_estadisticas.json")
            await ctx.send(f"📂 云盘模块：成功读取并解析了 {new_files_count} 个 Power BI 新报表，合计 {len(new_records)} 条宏观统计数据。")

    except HttpError as error:
        logging.error(f"Drive API Error: {error}")
        await ctx.send(f"Drive API 出现错误: {error}")

def parse_powerbi_excel(file_path):
    """Parses macro historical data from a Power BI Excel export."""
    records = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        
        # Mappings matching PowerBI_Estadisticas table
        mappings = {
            'fecha': ['fecha', 'date', '时间', '月', 'periodo', 'reporte'],
            'categoria': ['categoria', 'category', '类别', '大类'],
            'subcategoria': ['subcategoria', 'sub', '子类别', '小类'],
            'producto': ['producto', 'product', '产品', '商品', '名称'],
            'ventas': ['ventas', 'sales', '销售额', '金额'],
            'ventas_prev': ['anterior', 'anterior_ano', '去年同期', '同比', 'prev_year'],
            'n_total': ['total_articulos', 'total', '数量', '件数', 'n_total_articulos']
        }
        
        idx_map = {}
        for key, aliases in mappings.items():
            idx_map[key] = next((i for i, h in enumerate(headers) if any(a in h.lower() for a in aliases)), -1)

        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            
            fecha = row[idx_map['fecha']].value if idx_map['fecha'] != -1 else None
            categoria = row[idx_map['categoria']].value if idx_map['categoria'] != -1 else "未分类"
            subcat = row[idx_map['subcategoria']].value if idx_map['subcategoria'] != -1 else "未分类"
            prod = row[idx_map['producto']].value if idx_map['producto'] != -1 else None
            ventas = row[idx_map['ventas']].value if idx_map['ventas'] != -1 else 0
            ventas_prev = row[idx_map['ventas_prev']].value if idx_map['ventas_prev'] != -1 else 0
            n_total = row[idx_map['n_total']].value if idx_map['n_total'] != -1 else 0

            if not prod:
                continue

            records.append({
                "fecha_reporte": str(fecha).split(' ')[0] if fecha else datetime.now().strftime("%Y-%m-%d"),
                "categoria": str(categoria).strip(),
                "subcategoria": str(subcat).strip(),
                "producto": str(prod).strip(),
                "ventas": float(ventas) if ventas else 0.0,
                "ventas_ano_anterior": float(ventas_prev) if ventas_prev else 0.0,
                "n_total_articulos": int(n_total) if n_total else 0
            })
            
        wb.close()
    except Exception as e:
        logging.error(f"Error parsing Power BI Excel: {e}")
    return records

# Helper to save lists to JSON files
def save_records_to_json(records, file_path):
    existing = []
    if os.path.exists(file_path):
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            existing = []
            
    # Append new records
    existing.extend(records)
    
    # Save back
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

# =============================================================================
# Agent Config and Main Loop
# =============================================================================
# Set triggers to poll every 5 minutes (300 seconds)
gmail_trigger = every(300, check_gmail_orders)
drive_trigger = every(300, check_gdrive_reports)

agent_config = LocalAgentConfig(
    system_instructions=(
        "You are the Mainpaper Ingestion Orchestrator. "
        "Your duty is to run triggers to capture Excel sales orders from Gmail and "
        "Power BI summary reports from Google Drive. Notify the user whenever data is parsed."
    ),
    triggers=[gmail_trigger, drive_trigger]
)

async def main():
    print("--------------------------------------------------")
    print("Mainpaper Ingestion Orchestrator Trigger Starting...")
    print("Watching Gmail orders and Google Drive reports.")
    print("Logs are written to ingest_agent.log.")
    print("--------------------------------------------------")
    
    # Ensure token is ready before loop starts
    creds = get_credentials()
    if not creds:
        print("CRITICAL: Failed to load Google credentials. Please authorize the app.")
        return

    async with Agent(agent_config) as agent:
        # Keep running to let triggers fire
        while True:
            await asyncio.sleep(3600)  # Sleep 1h incrementally

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\nOrchestrator stopped by user.")
